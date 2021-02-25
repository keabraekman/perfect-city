import sys
import time
import requests
import os
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl import load_workbook
import pickle

import re

workbook = load_workbook('perfect-city.xlsx')
worksheet = workbook.active

states = ["Alaska", "Alabama", "Arkansas", "Arizona", 
"California", "Colorado", "Connecticut", "District-of-Columbia", 
"Delaware", "Florida", "Georgia", "Hawaii", "Iowa", "Idaho", 
"Illinois", "Indiana", "Kansas", "Kentucky", "Louisiana", "Massachusetts", 
"Maryland", "Maine", "Michigan", "Minnesota", "Missouri", "Mississippi", 
"Montana", "North-Carolina", "North-Dakota", "Nebraska", "New-Hampshire", 
"New-Jersey", "New-Mexico", "Nevada", "New-York", "Ohio", "Oklahoma", 
"Oregon", "Pennsylvania", "Rhode Island", "South-Carolina", 
"South-Dakota", "Tennessee", "Texas", "Utah", "Virginia", 
"Vermont", "Washington", "Wisconsin", "West-Virginia", "Wyoming"]

column_titles = ['City', 'State', 'White', 'Hispanic', 'Black', 'Asian', 
'American Indian', 'Other', 'Pacific Islander', 'Two or more races', 'Population', 
'Average Age', 'Median Income per Capita', 'Median Income per Household', 
'Unemployment', 'Poverty level', 'Trump', 'Biden', 'Average BMI', 'Overweight percentage',
'number of sunny days', 'percentage of women', 'Violent Crime', 'Population Density', 'tax rate']

def get_html(url):
    re = requests.get(url)
    print(type(re.text))
    return BeautifulSoup(re.text, 'html.parser')

def get_cities(state):
    print('\t Adding ', state)
    soup = get_html('https://www.city-data.com/city/' + state + '.html')
    cities = []
    for c in soup.find_all('tr', {'class': 'rB'}):
        cities.append(c.find_all('td')[1].find_all('a')[0].contents[0])
    return cities


def fill_headers(titles):
    column = 1
    for t in titles:
        worksheet.cell(row = 1, column = column, value = t)
        column += 1


def create_city_list():
    cities = []
    for s in states:
        to_add = get_cities(s)
        for c in to_add:
            cities.append(c.split(',')[0])
    return cities


def fill_cities_column(cities):
    row = 2
    for c in cities:
        worksheet.cell(row = row, column = 1, value = c.replace(' ', '-'))
        row += 1


def number_of_cities(state):
    return len(get_cities(state))


def find_first_empty_row(column):
    r = 1
    while(worksheet.cell(row = r, column = column).value is not None):
        r += 1
    return r


def fill_state():
    for s in states:
        number_of_rows = number_of_cities(s)
        first_empty_row = find_first_empty_row(2)
        for i in range(number_of_rows):
            worksheet.cell(row = first_empty_row+i, column = 2, value = s)
        workbook.save('perfect-city.xlsx')


def white(soup, row):
    return soup.find(text='White alone').parent.parent.find(string=re.compile('%'))

def fill_white():
    row = 2
    # if(worksheet.cell(row = row, column = 3).value is not None)
    while(worksheet.cell(row = row, column = 1).value is not None):
        row += 1
        if(worksheet.cell(row = row, column = 3).value is None):
            city = worksheet.cell(row = row, column = 1).value
            state = worksheet.cell(row = row, column = 2).value
            print('ADDING : ' + city + ', ' + state + '\t row = ', row)
            print(100*row/6849 , '% Done')
            url = 'https://www.city-data.com/city/' + city + '-' + state + '.html'
            worksheet.cell(row = row, column = 3).value = white(get_html(url), row)
            if(row % 10 == 0):
                print('saved spreadsheet')
                workbook.save('perfect-city.xlsx')




# print('fill_headers')
# fill_headers(column_titles)
# print('Cities')
# cities = create_city_list()
# print('fill_cities_column')
# fill_cities_column(cities)
# print('fill_state')
# fill_state()

# workbook.save('perfect-city.xlsx')



# DELETE WHITE Column
# for i in range(1, 10000):
#     worksheet.cell(row = i, column = 3).value = None
# workbook.save('perfect-city.xlsx')



# print('Adding whites')
# fill_white()




def get_str(url):
    re = requests.get(url)
    return re.text

path = '/Users/keabraekman/Documents/Personal/perfect-city'



# for s in states:
#     print('creating : ' + path + '/' + s)
#     os.mkdir(path + '/' + s)


row = 2
while(worksheet.cell(row = row, column = 1).value is not None):
    city = worksheet.cell(row = row, column = 1).value
    state = worksheet.cell(row = row, column = 2).value
    print('creating : ' + path + '/' + state + '/' + city)
    os.mkdir(path + '/' + state + '/' + city)
    row += 1








# detect the current working directory and print it
# path = os.getcwd()
# print ("The current working directory is %s" % path)


# workbook.save('perfect-city.xlsx')]
